#!/usr/bin/env python3
"""
The bid-library provider seam — Complete (Stage 4)'s reuse corpus + evidence register.

`docs/design/architecture.md` calls for one interface the app reads the bid
library through, with swappable backends (the same pattern as `src/sources.py`):
  - **LocalMirror (now)** — reads FWF's real exported bid store held locally
    (`knowledge/SharePoint Folder/Bids/02 Bid Library/Bid Library Tracker.xlsx`).
    The export is gitignored (client-confidential; never committed), so this reads
    it at runtime and degrades cleanly to "unavailable" when it isn't present
    (e.g. CI) rather than faking data.
  - **GraphSharePoint (later)** — the live MS Graph connection, dropped in behind
    the same `LibraryProvider` interface with no app changes.

This is NOT invented: the tracker's 10 category sheets and 9 columns are FWF's
real `Bid Library Tracker` (data-model.md §4b — the reuse corpus AND the
compliance-asset register). One honest wrinkle the real data forced: **expiry is a
first-class field** the data-model demands, but the tracker keeps it in free-text
`Notes` ("9001 Expires: 09/01/2026", "ISO 27001 Expires 31 Oct 2025"), not a
column. So the mirror *extracts* expiry from the text and surfaces it structured —
the "facts decay / an expired cert at bid time is the founding-class failure"
principle (knowledge/VERIFIED_FACTS.md) realised against the data as it actually is.

The retrieval side (`search`) is a lean keyword/tag match — enough to give
Complete's AI pre-fill real few-shot material + an evidence ledger today; a
smarter retriever can slot in behind the same call later.
"""
import datetime
import os
import re


def _openpyxl():
    """The openpyxl module, or None if it isn't installed. Declared in
    requirements.txt as a hard dependency of Complete (Stage 4); this indirection
    exists only so a missing install degrades to an *honest* 'unavailable' state
    (see `status()`) instead of a crash or a contradictory 'connected, 0 items'."""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        return None

# The category sheets in the tracker (data-model.md §4b). Text-tolerant: an
# unknown sheet is still read, this is just the canonical order for the UI.
LIBRARY_CATEGORIES = [
    "Company Credentials", "Capabilities", "Case Studies", "Team & Resources",
    "Social Value", "Commercial", "Delivery", "Templates", "Governance", "Branding",
]

# One LibraryItem's fields, mapped 1:1 from the tracker's 9 columns (+ derived
# `category` from the sheet name, + the derived expiry fields added on read).
LIBRARY_COLUMNS = [
    "item", "description", "status", "owner", "review_frequency",
    "assigned_to", "progress_status", "last_updated", "notes",
]

# Categories whose items double as compliance evidence (certs, policies, cover) —
# the ones the Complete evidence ledger surfaces with their expiry.
EVIDENCE_CATEGORIES = {"Company Credentials", "Governance", "Commercial"}

# An expiry this many days out (or nearer) is "expiring soon" — flagged amber, not
# yet a hard fail. Matches the spirit of the Manage pre-flight expiry check.
EXPIRING_SOON_DAYS = 90

_HERE = os.path.dirname(os.path.abspath(__file__))
# The real export lives at the repo root (one level up from src/), gitignored.
_DEFAULT_ROOT = os.path.normpath(os.path.join(_HERE, "..", "knowledge", "SharePoint Folder", "Bids"))
_TRACKER_REL = os.path.join("02 Bid Library", "Bid Library Tracker.xlsx")
_MASTER_REL = os.path.join("01 Bid Forms", "FOR006 Tender Response Master.xlsx")

# The FOR006 master's 18 columns → our ResponseItem keys, mapped by header text.
_MASTER_HEADER_MAP = {
    "customer document": "customer_document",
    "section": "section",
    "sub-section": "sub_section",
    "question reference number": "question_ref",
    "questions": "question_text",
    "type of question": "question_type",
    "weightings %": "weighting_pct",
    "word count": "word_count_limit",
    "actual words": "actual_words",
    "images permitted?": "images_permitted",
    "attachments permitted ?": "attachments_permitted",
    "tags": "tags",
    "supplier response": "supplier_response",
    "owner": "owner",
    "supporting person": "supporting_person",
    "reviewer": "reviewer",
    "target date": "target_date",
    "status": "status",
}


def _root():
    """The bid-store root. Overridable via BID_LIBRARY_ROOT so the mirror can point
    at a different local export without code changes."""
    return os.environ.get("BID_LIBRARY_ROOT", _DEFAULT_ROOT)


def tracker_path():
    return os.path.join(_root(), _TRACKER_REL)


def master_path():
    return os.path.join(_root(), _MASTER_REL)


# A minimal built-in question set, used only when the real FOR006 master isn't
# present (e.g. CI) so Complete still has a matrix to render — clearly generic, not
# a claim to be FWF's real questions.
_FALLBACK_QUESTIONS = [
    {"section": "Quality", "question_ref": "Q1", "question_text": "Describe your technical approach and methodology.",
     "question_type": "Text Response", "weighting_pct": "30", "word_count_limit": "750", "tags": "Capabilities"},
    {"section": "Quality", "question_ref": "Q2", "question_text": "Describe your approach to social value.",
     "question_type": "Text Response", "weighting_pct": "10", "word_count_limit": "500", "tags": "Social Value"},
    {"section": "Quality", "question_ref": "Q3", "question_text": "Describe your information security and data protection approach.",
     "question_type": "Text Response", "weighting_pct": "20", "word_count_limit": "750", "tags": "Company Credentials"},
]


def master_template():
    """The FOR006 tender-response question set, read from the real master template
    if present, else a small generic fallback. Returns ResponseItem-shaped dicts —
    the seed for a bid's compliance matrix before any answers are written. Reading
    the real file keeps the matrix grounded in FWF's actual question structure."""
    path = master_path()
    openpyxl = _openpyxl()
    if not os.path.exists(path) or openpyxl is None:
        return [dict(q) for q in _FALLBACK_QUESTIONS]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(min_row=1, values_only=True)
    try:
        header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
    except StopIteration:
        wb.close()
        return [dict(q) for q in _FALLBACK_QUESTIONS]
    idx = {}
    for pos, name in enumerate(header):
        key = _MASTER_HEADER_MAP.get(name)
        if key and key not in idx:
            idx[key] = pos
    out = []
    for raw in rows:
        if not any(c not in (None, "") for c in raw):
            continue
        item = {key: _clean(raw[pos]) if pos < len(raw) else "" for key, pos in idx.items()}
        if not (item.get("question_ref") or item.get("question_text")):
            continue
        # A fresh matrix starts un-worked: drop any template answer/status.
        item["supplier_response"] = ""
        item["status"] = "To do"
        out.append(item)
    wb.close()
    return out or [dict(q) for q in _FALLBACK_QUESTIONS]


# --- Expiry extraction from free text ---------------------------------------
# The tracker stores renewal dates inside Notes/Item text, in mixed formats. These
# patterns pull them out; _to_iso normalises to YYYY-MM-DD so the UI + alerts can
# compare against today.
_MONTHS = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 1)}

# "09/01/2026", "31/10/25" (dd/mm/yyyy or dd/mm/yy)
_RE_DMY = re.compile(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b")
# "31 Oct 2025", "9 January 2026"
_RE_TEXT = re.compile(r"\b(\d{1,2})\s+([A-Za-z]{3,9})\s+(\d{4})\b")
# ISO "2026-01-09"
_RE_ISO = re.compile(r"\b(\d{4})-(\d{2})-(\d{2})\b")


def _to_iso(y, m, d):
    """Validate + normalise a (year, month, day) to an ISO date string, or None."""
    if y < 100:
        y += 2000
    try:
        return datetime.date(y, m, d).isoformat()
    except ValueError:
        return None


def extract_expiries(text):
    """Every expiry-looking date in a blob of text, as ISO strings (sorted, unique).
    Only counts dates that sit near an expiry/renewal cue, so a 'Last updated' date
    in prose isn't mistaken for an expiry."""
    if not text:
        return []
    s = str(text)
    low = s.lower()
    # Only mine for expiry if the text signals one — keeps false positives down.
    if not any(cue in low for cue in ("expir", "renew", "valid until", "expires", "due")):
        return []
    found = set()
    for m in _RE_DMY.finditer(s):
        iso = _to_iso(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        if iso:
            found.add(iso)
    for m in _RE_ISO.finditer(s):
        iso = _to_iso(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        if iso:
            found.add(iso)
    for m in _RE_TEXT.finditer(s):
        mon = _MONTHS.get(m.group(2).lower()[:3])
        if mon:
            iso = _to_iso(int(m.group(3)), mon, int(m.group(1)))
            if iso:
                found.add(iso)
    return sorted(found)


def _days_until(iso_date, now=None):
    try:
        d = datetime.date.fromisoformat(iso_date)
    except (TypeError, ValueError):
        return None
    today = (now or datetime.datetime.now(datetime.timezone.utc)).date()
    return (d - today).days


def _expiry_fields(item, now=None):
    """Add derived expiry to a LibraryItem: the soonest extracted expiry + a
    status (expired / expiring_soon / ok / none)."""
    dates = extract_expiries(f"{item.get('item', '')} {item.get('notes', '')}")
    item["expiries"] = dates
    if not dates:
        item["expiry_date"] = None
        item["expiry_status"] = "none"
        item["days_to_expiry"] = None
        return item
    soonest = dates[0]
    days = _days_until(soonest, now)
    item["expiry_date"] = soonest
    item["days_to_expiry"] = days
    if days is not None and days < 0:
        item["expiry_status"] = "expired"
    elif days is not None and days <= EXPIRING_SOON_DAYS:
        item["expiry_status"] = "expiring_soon"
    else:
        item["expiry_status"] = "ok"
    return item


# --- Providers --------------------------------------------------------------

class LibraryProvider:
    """The seam. A backend supplies `items()`; everything else (search, evidence,
    status) is computed on top so a new backend only implements the read."""
    name = "base"

    def items(self):
        """All LibraryItems as dicts (category + LIBRARY_COLUMNS + derived expiry)."""
        raise NotImplementedError

    def available(self):
        return True

    def unavailable_reason(self):
        """Human-readable reason `available()` is False, or None. Surfaced in
        `status()` so a misconfigured deployment shows *why* the library is dark
        instead of a silent, contradictory 'connected, 0 items'."""
        return None

    def status(self):
        avail = self.available()
        st = {
            "provider": self.name,
            "available": avail,
            "source": self.source_label(),
            "count": len(self.items()) if avail else 0,
        }
        if not avail:
            st["reason"] = self.unavailable_reason()
        return st

    def source_label(self):
        return self.name


class LocalMirrorProvider(LibraryProvider):
    """Reads FWF's real exported Bid Library Tracker from the local store. Never
    writes; never fakes — if the file isn't present, `available()` is False and the
    app shows an honest 'library not connected' state."""
    name = "local_mirror"

    def __init__(self):
        # The API constructs a fresh provider per request, so an instance-level memo
        # is safe and stops status()+items() from parsing the workbook twice a call.
        self._items_cache = None

    def available(self):
        # Both must hold: the export must exist AND openpyxl must be installed to
        # read it. Without the latter, items() would return [] while the file is
        # present — an honest 'unavailable' beats a contradictory 'connected, 0'.
        return os.path.exists(tracker_path()) and _openpyxl() is not None

    def unavailable_reason(self):
        if not os.path.exists(tracker_path()):
            return "Bid Library Tracker.xlsx not found at the configured path"
        if _openpyxl() is None:
            return "openpyxl is not installed (pip install openpyxl) — the tracker can't be read"
        return None

    def source_label(self):
        return "LocalMirror · Bid Library Tracker.xlsx"

    def items(self):
        if self._items_cache is not None:
            return self._items_cache
        openpyxl = _openpyxl()
        if not os.path.exists(tracker_path()) or openpyxl is None:
            self._items_cache = []
            return self._items_cache
        wb = openpyxl.load_workbook(tracker_path(), read_only=True, data_only=True)
        out = []
        for ws in wb.worksheets:
            category = ws.title
            rows = ws.iter_rows(min_row=1, values_only=True)
            try:
                header = [str(c).strip() if c is not None else "" for c in next(rows)]
            except StopIteration:
                continue
            # Map the sheet's header names to our column keys, tolerating wording.
            idx = _column_index(header)
            for raw in rows:
                if not any(c not in (None, "") for c in raw):
                    continue
                item = {"category": category}
                for key, i in idx.items():
                    val = raw[i] if i is not None and i < len(raw) else None
                    item[key] = _clean(val)
                if not item.get("item"):
                    continue  # a row with no Item is a spacer, skip it
                out.append(_expiry_fields(item))
        wb.close()
        self._items_cache = out
        return out


# Header wording → our column keys. The tracker uses "Team & Resources",
# "Progress Status", etc.; map by normalised header text.
_HEADER_MAP = {
    "item": "item",
    "description": "description",
    "status": "status",
    "owner": "owner",
    "review frequency": "review_frequency",
    "assigned to": "assigned_to",
    "progress status": "progress_status",
    "last updated": "last_updated",
    "notes": "notes",
}


def _column_index(header):
    """{column_key: position} for the columns we recognise in this sheet's header."""
    idx = {k: None for k in LIBRARY_COLUMNS}
    for pos, name in enumerate(header):
        key = _HEADER_MAP.get(str(name).strip().lower())
        if key and idx.get(key) is None:
            idx[key] = pos
    return idx


def _clean(val):
    """Cell → tidy string. Dates render as ISO; 'None'/'N/A' placeholders blank out."""
    if val in (None, ""):
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.date().isoformat() if isinstance(val, datetime.datetime) else val.isoformat()
    s = str(val).strip()
    return "" if s.lower() in ("none", "n/a", "na") else s


# --- Retrieval + evidence (computed on top of any provider) -----------------

_STOPWORDS = {"the", "a", "an", "of", "and", "or", "to", "for", "in", "on", "with",
              "how", "do", "you", "your", "our", "we", "will", "be", "is", "are",
              "please", "describe", "provide", "detail", "outline", "explain"}


def _tokens(text):
    return {w for w in re.findall(r"[a-z0-9]+", (text or "").lower())
            if len(w) > 2 and w not in _STOPWORDS}


def search(items, query="", tags="", limit=5):
    """Rank LibraryItems against a question + its tags — the few-shot / evidence
    retrieval Complete's AI pre-fill draws from. A lean keyword overlap over the
    item, description, notes, tags and category; enough to surface real material
    today, swappable for a smarter retriever later. Returns the top `limit` with a
    `score` each."""
    q = _tokens(query) | _tokens(tags)
    scored = []
    for it in items:
        hay = _tokens(" ".join(str(it.get(f, "")) for f in
                               ("item", "description", "notes", "category")))
        overlap = len(q & hay)
        # A tag/category match is worth more than an incidental word hit.
        cat_bonus = 2 if _tokens(it.get("category", "")) & q else 0
        score = overlap + cat_bonus
        if score > 0:
            scored.append((score, it))
    scored.sort(key=lambda s: s[0], reverse=True)
    return [{**it, "score": score} for score, it in scored[:limit]]


def evidence(items, now=None):
    """The compliance-evidence ledger: items that carry an expiry, or that live in
    the credential/governance/commercial categories — surfaced soonest-expiry-first
    so a cert about to lapse is impossible to miss at bid time."""
    led = [it for it in items
           if it.get("expiry_date") or it.get("category") in EVIDENCE_CATEGORIES]
    # Expired/soon first, then dated, then the rest.
    rank = {"expired": 0, "expiring_soon": 1, "ok": 2, "none": 3}
    led.sort(key=lambda it: (rank.get(it.get("expiry_status"), 3),
                             it.get("days_to_expiry") if it.get("days_to_expiry") is not None else 10**6))
    return led


def get_provider():
    """The active library provider (LocalMirror now; GraphSharePoint later via
    LIBRARY_PROVIDER, mirroring llm.get_provider / sources)."""
    key = os.environ.get("LIBRARY_PROVIDER", "local_mirror").strip().lower()
    if key == "local_mirror":
        return LocalMirrorProvider()
    # GraphSharePoint slots in here behind the same interface — not built (no MS
    # Graph in this environment; CLAUDE.md hard rule).
    if key == "graph_sharepoint":
        raise RuntimeError(
            "LIBRARY_PROVIDER=graph_sharepoint is not built yet (no MS Graph access). "
            "Use local_mirror until GraphSharePointProvider lands.")
    # Any other value is a config typo — fail loudly rather than silently reading
    # the local mirror and reporting 'not connected' with no hint why (mirrors
    # llm.get_provider()).
    raise RuntimeError(
        f"Unknown LIBRARY_PROVIDER '{key}'. Valid: local_mirror (graph_sharepoint planned).")


def reference():
    """Library vocabulary for the UI (categories, expiry window)."""
    return {
        "categories": LIBRARY_CATEGORIES,
        "evidence_categories": sorted(EVIDENCE_CATEGORIES),
        "expiring_soon_days": EXPIRING_SOON_DAYS,
    }


if __name__ == "__main__":
    prov = get_provider()
    st = prov.status()
    print(f"Library provider: {st['provider']} · available={st['available']} · {st['source']}")
    if not st["available"]:
        print(f"  (tracker not found at {tracker_path()})")
    else:
        items = prov.items()
        print(f"  {len(items)} library items across "
              f"{len({i['category'] for i in items})} categories")
        exp = [i for i in items if i["expiry_status"] in ("expired", "expiring_soon")]
        print(f"  {len(exp)} item(s) expired / expiring within {EXPIRING_SOON_DAYS}d:")
        for i in exp:
            print(f"    · [{i['category']}] {i['item']} — {i['expiry_status']} "
                  f"({i['expiry_date']}, {i['days_to_expiry']}d)")
