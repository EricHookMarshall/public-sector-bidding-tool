#!/usr/bin/env python3
"""B02 — Compliance Matrix builder (v2).

Builds the full-traceability requirements matrix with FIVE separate RAG
statuses (compliance / evidence / drafting / review / submission), so a row
that is drafted but not evidenced does not read as green.

Extraction is Claude's job; this script builds the workbook deterministically.

Usage:
    python build_matrix.py --in requirements.json --out matrix.xlsx --title "..."

Each requirement object may contain any of the columns in COLUMNS (below).
RAG fields default to "Red".
"""
import argparse
import json
import sys

COLUMNS = [
    ("ref", "Ref"),
    ("source_doc", "Source document"),
    ("locator", "Page/para/clause"),
    ("requirement", "Requirement"),
    ("category", "Category"),
    ("type", "M/D"),
    ("pass_fail", "Pass/fail?"),
    ("answer_required", "Answer req'd?"),
    ("eval_criterion", "Evaluation criterion"),
    ("weighting", "Weight/score"),
    ("limit", "Limit"),
    ("lot", "Lot"),
    ("draft_owner", "Draft owner"),
    ("evidence_owner", "Evidence owner"),
    ("dependency", "Dependency"),
    ("draft_link", "Draft link"),
    ("evidence_link", "Evidence link"),
    ("risk", "Risk"),
    ("clarification", "Clarify?"),
    ("submission_location", "Submission location"),
    ("rag_compliance", "RAG Compliance"),
    ("rag_evidence", "RAG Evidence"),
    ("rag_drafting", "RAG Drafting"),
    ("rag_review", "RAG Review"),
    ("rag_submission", "RAG Submission"),
    ("notes", "Notes"),
]

RAG_COLS = ["rag_compliance", "rag_evidence", "rag_drafting", "rag_review",
            "rag_submission"]
RAG_FILL = {"Red": "F4CCCC", "Amber": "FCE5CD", "Green": "D9EAD3"}


def build(reqs, out_path, title):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("openpyxl not installed. pip install openpyxl --break-system-packages")

    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Matrix"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    header_fill = PatternFill("solid", fgColor="1F2A44")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, (_, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=c, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border

    for r, req in enumerate(reqs, start=3):
        for c, (key, _) in enumerate(COLUMNS, start=1):
            if key in RAG_COLS:
                val = (req.get(key) or "Red").capitalize()
            else:
                val = req.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if key in RAG_COLS and val in RAG_FILL:
                cell.fill = PatternFill("solid", fgColor=RAG_FILL[val])

    widths = {"Requirement": 55, "Evaluation criterion": 30, "Source document": 20,
              "Submission location": 22, "Notes": 22, "Dependency": 14,
              "Draft link": 18, "Evidence link": 18}
    for c, (_, label) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(label, 12)

    ws.freeze_panes = "D3"
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="Build compliance matrix v2")
    ap.add_argument("--in", dest="infile", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--title", default="Compliance Matrix")
    args = ap.parse_args()
    with open(args.infile, encoding="utf-8") as fh:
        reqs = json.load(fh)
    build(reqs, args.out, args.title)

    mand = [r for r in reqs if (r.get("type") or "").upper() == "M"]
    not_green = lambda r: any((r.get(c) or "Red").capitalize() != "Green"
                              for c in RAG_COLS)
    print(f"Wrote {len(reqs)} requirements to {args.out}")
    print(f"  Mandatory: {len(mand)}")
    print(f"  Mandatory rows not fully Green (all 5 RAGs): "
          f"{sum(1 for r in mand if not_green(r))}")
    print(f"  Requirements flagged for clarification: "
          f"{sum(1 for r in reqs if str(r.get('clarification','')).lower() in ('yes','true','y'))}")


if __name__ == "__main__":
    main()
